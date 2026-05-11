from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Market
from app.services.risk_calibration import risk_calibration_for_market
from app.services.risk_engine import RiskInputs, ScenarioSpec, assess_risk


def _path_fan_svg(price_paths: list[list[float]], width: int = 720, height: int = 240) -> str:
    if not price_paths:
        return f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" />'
    values = [value for path in price_paths for value in path]
    lo = min(values)
    hi = max(values)
    span = max(1e-9, hi - lo)
    longest = max(len(path) for path in price_paths)
    denom = max(1, longest - 1)
    polylines = []
    for path in price_paths[:200]:
        points = []
        for idx, value in enumerate(path):
            x = (idx / denom) * width
            y = height - ((value - lo) / span) * height
            points.append(f"{x:.1f},{y:.1f}")
        polylines.append(
            f'<polyline points="{" ".join(points)}" fill="none" stroke="#2563eb" '
            'stroke-width="1" stroke-opacity="0.08" />'
        )
    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
        f'viewBox="0 0 {width} {height}">'
        '<rect width="100%" height="100%" fill="#ffffff" />'
        + "".join(polylines)
        + "</svg>"
    )


def _assessment_for_export(db: Session, payload: Any) -> dict[str, Any]:
    return assess_risk(
        db,
        RiskInputs(
            market_code=payload.market_code,
            position_gbp=payload.position_gbp,
            position_unit=payload.position_unit,
            position_mwh=payload.position_mwh,
            hedge_ratio=payload.hedge_ratio,
            horizon_hours=payload.horizon_hours,
            target_timestamp=payload.target_timestamp,
            direction=payload.direction,
            n_paths=payload.n_paths,
            scenarios=[
                ScenarioSpec(
                    name=s.name,
                    sigma_multiplier=s.sigma_multiplier,
                    drift_shift=s.drift_shift,
                    spot_shock_pct=s.spot_shock_pct,
                )
                for s in payload.scenarios
            ],
            basis_against_market_code=payload.basis_against_market_code,
            basis_direction=payload.basis_direction,
            random_seed=260514,
            path_sample_size=200,
        ),
    )


def _pack_payload(db: Session, payload: Any) -> dict[str, Any]:
    assessment = _assessment_for_export(db, payload)
    market = db.scalar(select(Market).where(Market.code == assessment["market_code"]))
    calibration = risk_calibration_for_market(db, market.id) if market else None
    path_fan_svg = _path_fan_svg(assessment.get("price_paths", []))
    return {
        "assessment": assessment,
        "calibration": calibration,
        "path_fan_svg": path_fan_svg,
        "fx": {
            "price_currency": assessment.get("price_currency"),
            "fx_to_gbp": assessment.get("fx_to_gbp"),
        },
    }


def _summary_rows(pack: dict[str, Any]) -> list[tuple[str, Any]]:
    assessment = pack["assessment"]
    return [
        ("Timestamp", assessment["as_of"]),
        ("Market", f"{assessment['market_name']} ({assessment['market_code']})"),
        ("Direction", assessment["direction"]),
        ("Horizon hours", assessment["horizon_hours"]),
        ("Position GBP", assessment["position_gbp"]),
        ("Risk GBP", assessment["risk_gbp"]),
        ("Likely GBP", assessment["likely_gbp"]),
        ("Upside GBP", assessment["upside_gbp"]),
        ("Spot price", assessment["spot_price"]),
        ("Forecast price", assessment["forecast_price"]),
        ("FX to GBP", assessment["fx_to_gbp"]),
        ("Price currency", assessment["price_currency"]),
        ("Regime", assessment["regime"]),
        ("Scorer provider", assessment["scorer_provider"]),
    ]


def _xlsx_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def build_risk_export(db: Session, payload: Any, export_format: str) -> tuple[bytes, str, str, dict[str, Any]]:
    pack = _pack_payload(db, payload)
    assessment = pack["assessment"]
    filename = f"risk_assessment_{assessment['market_code']}_{assessment['as_of'].strftime('%Y%m%d_%H%M%S')}.{export_format}"
    if export_format == "pdf":
        content = _build_pdf(pack)
        media_type = "application/pdf"
    elif export_format == "xlsx":
        content = _build_xlsx(pack)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        raise ValueError("unsupported export format")
    audit_payload = {
        "format": export_format,
        "market_code": assessment["market_code"],
        "risk_gbp": assessment["risk_gbp"],
        "likely_gbp": assessment["likely_gbp"],
        "upside_gbp": assessment["upside_gbp"],
    }
    return content, media_type, filename, audit_payload


def _build_pdf(pack: dict[str, Any]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Risk assessment export")
    styles = getSampleStyleSheet()
    assessment = pack["assessment"]
    story = [
        Paragraph("Risk Assessment Export", styles["Title"]),
        Paragraph(assessment["rationale"], styles["BodyText"]),
        Spacer(1, 12),
        Table([["Field", "Value"], *[(key, str(value)) for key, value in _summary_rows(pack)]]),
        Spacer(1, 12),
        Paragraph("Coefficients", styles["Heading2"]),
    ]
    coefficient_rows = [["Key", "Label", "Value", "Unit", "Group"]]
    for item in assessment.get("coefficients", {}).get("items", []):
        coefficient_rows.append([item["key"], item["label"], item["value"], item["unit"], item["group"]])
    coefficient_table = Table(coefficient_rows, repeatRows=1)
    coefficient_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.append(coefficient_table)
    story.extend(
        [
            Spacer(1, 12),
            Paragraph("Calibration", styles["Heading2"]),
            Paragraph(str(pack["calibration"] or {}), styles["Code"]),
            Spacer(1, 12),
            Paragraph("Path Fan SVG", styles["Heading2"]),
            Paragraph(pack["path_fan_svg"].replace("<", "&lt;").replace(">", "&gt;")[:3500], styles["Code"]),
        ]
    )
    for flowable in story:
        if isinstance(flowable, Table):
            flowable.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                    ]
                )
            )
    doc.build(story)
    return buffer.getvalue()


def _build_xlsx(pack: dict[str, Any]) -> bytes:
    from openpyxl import Workbook

    buffer = BytesIO()
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    for row in _summary_rows(pack):
        summary.append([_xlsx_value(value) for value in row])

    coefficients = workbook.create_sheet("Coefficients")
    coefficients.append(["key", "label", "value", "unit", "group", "description"])
    for item in pack["assessment"].get("coefficients", {}).get("items", []):
        coefficients.append([item["key"], item["label"], item["value"], item["unit"], item["group"], item["description"]])

    scenarios = workbook.create_sheet("Scenarios")
    scenarios.append(["name", "risk_gbp", "likely_gbp", "upside_gbp", "prob_loss"])
    for item in pack["assessment"].get("scenarios", []):
        scenarios.append([item["name"], item["risk_gbp"], item["likely_gbp"], item["upside_gbp"], item["prob_loss"]])

    calibration = workbook.create_sheet("Calibration")
    for key, value in (pack["calibration"] or {}).items():
        calibration.append([key, _xlsx_value(value)])
    for key, value in pack["fx"].items():
        calibration.append([key, value])

    svg = workbook.create_sheet("PathFanSVG")
    svg["A1"] = pack["path_fan_svg"]
    workbook.save(buffer)
    return buffer.getvalue()
